from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Font, numbers, Alignment
import pandas as pd
from app import db
from app.models import Project
from datetime import date
    
def generate_app_xlsx(projects):
    # janky but will have to do
    workbook = load_workbook("tmp/APP_template.xlsx")
    workbook.iso_dates = True
    sheet = workbook.active
    year = date.today().year + 1
    
    table_title = ["", f"Bohol Island State University - Cogtong Candijay Campus Annual Procurement Plan for FY {year}"]
    table_header = ["Code (PAP)", "Procurement Project", "PMO/End-User", "Category", "Mode of Procurement", "Date Needed", "Source of Funds", "Estimated Budget", "Remarks"]
    sheet.append(table_title)
    sheet.append(table_header)
    
    count = 0
    for project in projects:
        count += 1
        project_to_append = [f'P{count}', project.title, project.submitter.department, project.category, project.initial_mode, project.date_needed.date(), project.source, project.budget, project.description]
        sheet.append(project_to_append)

    
    table_total = ["", "", "", "", "", "", "TOTAL:", f'=SUM(H3:H{count + 2})']
    sheet.append(table_total)
    
    sheet[f'B{count + 6}'] = "Prepared by:"
    sheet[f'E{count + 6}'] = "Approved by:"
    
    sheet['B1'].font = Font(name='Calibri', bold=True)
    set_header(sheet, 'A2:I2')
    set_border(sheet, f'A2:I{count + 3}') 
    set_header_center_alignment(sheet, 'A2:I2')
    set_vertical_center_alignment(sheet, 'A3:I1000')
    set_separator_format(sheet, 'H3:H1000')
    set_all_center_alignment(sheet, 'F3:H1000')
    set_wrap_text(sheet, 'B3:B1000')
    set_wrap_text(sheet, 'I3:I1000')
    
    workbook.save("tmp/APP_generated.xlsx")
    
def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            
def set_header(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.font = Font(name='Calibri', bold=True)

def set_separator_format(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            
def set_wrap_text(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            
def set_header_center_alignment(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def set_vertical_center_alignment(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.alignment = Alignment(vertical="center")
            
def set_all_center_alignment(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)